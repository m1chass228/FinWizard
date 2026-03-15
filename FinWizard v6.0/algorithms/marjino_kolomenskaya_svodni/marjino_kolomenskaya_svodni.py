import ast 
import string
import sys
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import column_index_from_string

line = sys.argv[1]   #  принимаем строку из c++
output_path = sys.argv[2]

file_names = ast.literal_eval(line)

output_wb = Workbook()
output_ws = output_wb.active
output_ws.title = "МарьиноКоломенскаяОтчет"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")
header_align = Alignment(horizontal="center", vertical="center")
thin = Side(border_style="thin", color="000000")
header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

columns_fill = PatternFill("solid", fgColor="E0FFFF")
columns_align = Alignment(horizontal="center", vertical="center")
def style_columns(cell):
    #cell.fill = columns_fill
    cell.alignment = columns_align

# ------------------------
# Заголовки
# ------------------------
headers = [
    "Дата", "Сотрудники", "Всего", "Вет.услуги",
    "Дерматолог", "Ратолог", "Стоматолог", "Кардиолог",
    "Орнитолог", "Ортопед", "Сумма з/п УС (пока нету)",
    "Анализы", "Ритуал", "Аптека", "Кол-во счетов",
    "Кол-во чеков", "Неовет", "Веттест"
]

for col, title in enumerate(headers, start=1):
    cell = output_ws.cell(row=1, column=col, value=title)
    style_header(cell)

# ------------------------
# Ширины колонок
# ------------------------
column_widths = {
    1: 12,   # Дата
    2: 40,   # Сотрудники
    3: 15,   # Всего
    4: 15,   # Вет.услуги
    5: 15,
    6: 15,
    7: 15,
    8: 15,
    9: 15,
    10: 15,
    11: 15,
    12: 15,
    13: 15,
    14: 15,
    15: 15,
    16: 15,
    17: 15, 
    18: 15
}


for col, width in column_widths.items():
    output_ws.column_dimensions[chr(64 + col)].width = width

rows_indent = 2
def print_values(cell_address, col_out, input_ws, output_ws, row_out):
    try:
        row_in = int(''.join(filter(str.isdigit, cell_address)))  # достаем цифры
        col_in = column_index_from_string(''.join(filter(str.isalpha, cell_address)))  # достаем буквы
        value = input_ws.cell(row=row_in, column=col_in).value
        if value is None:
            return
        value = float(value)
        cell = output_ws.cell(row=row_out, column=col_out, value=value)
        style_columns(cell)
        print(value)
    except Exception as e:
        print(f"[ОШИБКА] Не удалось прочитать {cell_address} в файле {filename}: {e}")

for filename in file_names:
    try:
        input_wb = openpyxl.load_workbook(filename, data_only=True)
    except Exception as e:
        print(f"[ОШИБКА] Не удалось открыть файл {filename}: {e}")
        continue
    input_ws = input_wb.active  #активный лист 

    # Достаем дату
    try:
        value = input_ws["E1"].value
        value = str(value).rstrip(".")
        cell = output_ws.cell(row=rows_indent, column=1, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    except Exception as e:
        print(f"[ОШИБКА] Не удалось прочитать E1 в файле {filename}: {e}")
        continue
    
    #достаем сотрудники
    allnames_list = []
    try:  
        for row in range(6, 8):
            for col in range(1, 4):
                value = input_ws.cell(row=row, column=col).value
                if value and str(value).strip():
                    allnames_list.append(str(value).strip())
        allnames = ", ".join(allnames_list)
        cell = output_ws.cell(row=rows_indent, column=2, value=allnames)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    except Exception as e:
        print(f"[ОШИБКА] Не удалось прочитать E1 в файле {filename}: {e}")
        continue
    # достаем Всего
    print_values("D10", 3, input_ws, output_ws, rows_indent)
    # достаем Вет.Услуги
    print_values("A11", 4, input_ws, output_ws, rows_indent)
    # достаем Дерматолог
    print_values("D18", 5, input_ws, output_ws, rows_indent)
    # достаем Ратолог
    print_values("D16", 6, input_ws, output_ws, rows_indent)
    # достаем Стоматолог
    print_values("D13", 7, input_ws, output_ws, rows_indent)
    # достаем Кардиолог
    print_values("D15", 8, input_ws, output_ws, rows_indent)
    # достаем Орнитолог
    print_values("D19", 9, input_ws, output_ws, rows_indent)
    # достаем Ортопед
    print_values("D17", 10, input_ws, output_ws, rows_indent)
    # сумма з/п УС
    # достаем Анализы
    analyses = 0.0
    try:
        value = input_ws["G19"].value
        if value is not None: 
            value = float(value)
            analyses += value 
        value = input_ws["G20"].value
        if value is not None: 
            value = float(value)
            analyses += value 
        cell = output_ws.cell(row=rows_indent, column=12, value=analyses)
        style_columns(cell)
    except Exception as e:
        print(f"[ОШИБКА] Не удалось прочитать G19 или G20 в файле {filename}: {e}")
    # достаем Ритуал 
    print_values("G21", 13, input_ws, output_ws, rows_indent)
    # достаем Аптека
    print_values("C10", 14, input_ws, output_ws, rows_indent)
    # достаем кол-во счетов
    print_values("H59", 15, input_ws, output_ws, rows_indent)
    # достаем кол-во чеков
    print_values("J59", 16, input_ws, output_ws, rows_indent)
    # достаем Неовет
    print_values("G19", 17, input_ws, output_ws, rows_indent)
    # достаем Веттест
    print_values("G20", 18, input_ws, output_ws, rows_indent)

    rows_indent += 1
try:
    output_wb.save(output_path)
except:
    print("ошибка записи")



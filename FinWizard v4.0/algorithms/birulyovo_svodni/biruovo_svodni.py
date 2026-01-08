import ast 
import sys
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import column_index_from_string


print("python запущен")
line = sys.argv[1]   #  принимаем строку из c++
output_path = sys.argv[2]

file_names = ast.literal_eval(line)

output_wb = Workbook()
output_ws = output_wb.active
output_ws.title = "БирюлевоОтчет"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")
header_align = Alignment(horizontal="center", vertical="center")
thin = Side(border_style="thin", color="000000")
header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

columns_fill = PatternFill("solid", fgColor="E0FFFF")
columns_align = Alignment(horizontal="center", vertical="center")
def style_columns(cell):
    #cell.fill = columns_fill
    cell.alignment = columns_align

# ------------------------
# Заголовки
# ------------------------
headers = [
    "Дата", "Смена", "Выручка за смену", "Услуги/день",
    "Услуги ночь", "Услуги всего", "З/п врачей/всего", "З/п 10%",
    "Ритуал", "Митрохина", "Ратолог",
    "Дерматолог", "Стоматолог", "Белозор", "Онколог", "Кардиолог",
    "Чеки день", "Чеки ночь", "Чеки аптека", "Аптека"
]

for col, title in enumerate(headers, start=1):
    cell = output_ws.cell(row=1, column=col, value=title)
    style_header(cell)

# ------------------------
# Ширины колонок
# ------------------------
column_widths = {
    1: 12,   # Дата
    2: 70,   # Смена
    3: 35,   # Выручка за смену
    4: 15,   # услуги/день
    5: 15,   # услуги/ночь
    6: 15,   # Услуги всего
    7: 15,   # З/п врачей
    8: 15,   # з/п 10
    9: 15,   # ритуал
    10: 15,  # митрохина
    11: 15,  # Ратолог
    12: 15,  # Дерматолог
    13: 15,  # Стоматолог
    14: 15,  # белозор 
    15: 15,  # Онколог 
    16: 15,  # Кардиолог 
    17: 15,  # чеки день
    18: 15,  # чеки ночь
    19: 15,  # чеки аптека
    20: 15,  # аптека

}


for col, width in column_widths.items():
    output_ws.column_dimensions[chr(64 + col)].width = width

rows_indent = 2
def print_values(cell_address, col_out, input_ws, output_ws, row_out, filename):
    try:
        row_in = int(''.join(filter(str.isdigit, cell_address)))  # достаем цифры
        col_in = column_index_from_string(''.join(filter(str.isalpha, cell_address)))  # достаем буквы
        value = input_ws.cell(row=row_in, column=col_in).value
        if value is None:
            return
        value = float(value)
        cell = output_ws.cell(row=row_out, column=col_out, value=value)
        style_columns(cell)
    except Exception as e:
        print(f"[ОШИБКА] Не удалось прочитать {cell_address} в файле {filename}: {e}")

for filename in file_names:
    try:
        input_wb = openpyxl.load_workbook(filename, data_only=True)
    except Exception as e:
        print(f"[ОШИБКА] Не удалось открыть файл {filename}: {e}")
        continue
    input_ws = input_wb.active  #активный лист 

    # Достаем дату
    try:
        value = input_ws["E1"].value
        value = str(value).rstrip(".")
        cell = output_ws.cell(row=rows_indent, column=1, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    except Exception as e:
        print(f"[ОШИБКА] Не удалось прочитать E1 в файле {filename}: {e}")
        continue
    
    #достаем смена
    allnames_list = []
    try:  
        for row in range(6, 8):
            for col in range(1, 4):
                value = input_ws.cell(row=row, column=col).value
                if value and str(value).strip():
                    allnames_list.append(str(value).strip())
        allnames = ", ".join(allnames_list)
        cell = output_ws.cell(row=rows_indent, column=2, value=allnames)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    except Exception as e:
        print(f"[ОШИБКА] Не удалось прочитать E1 в файле {filename}: {e}")
        continue
    # достаем Выручка за смену
    print_values("D13", 3, input_ws, output_ws, rows_indent, filename)
    # достаем Услуги/день
    print_values("A13", 4, input_ws, output_ws, rows_indent, filename)
    # достаем Услуги/ночь
    print_values("B13", 5, input_ws, output_ws, rows_indent, filename)
    # достаем Услуги всего
    print_values("A14", 6, input_ws, output_ws, rows_indent, filename)
    # достаем З/п врачей
    print_values("C10", 7, input_ws, output_ws, rows_indent, filename)
    # достаем З/п 10%
    print_values("D36", 8, input_ws, output_ws, rows_indent, filename)
    # достаем Ритуал
    print_values("G25", 9, input_ws, output_ws, rows_indent, filename)
    # достаем Митрохина
    print_values("G22", 10, input_ws, output_ws, rows_indent, filename)
    # достаем Ратолог
    print_values("D18", 11, input_ws, output_ws, rows_indent, filename)
    # достаем Дерматолог
    print_values("D19", 12, input_ws, output_ws, rows_indent, filename)
    # достаем Стоматолог
    print_values("B33", 13, input_ws, output_ws, rows_indent, filename)
    # достаем Белозор
    print_values("D16", 14, input_ws, output_ws, rows_indent, filename)
    # достаем Онколог
    print_values("B34", 15, input_ws, output_ws, rows_indent, filename)
    # достаем Кардиолог
    print_values("D17", 16, input_ws, output_ws, rows_indent, filename)
    # достаем Чеки день
    print_values("E65", 17, input_ws, output_ws, rows_indent, filename)
    # достаем Чеки ночь
    print_values("G65", 18, input_ws, output_ws, rows_indent, filename)
    # достаем Чеки аптека
    print_values("I65", 19, input_ws, output_ws, rows_indent, filename)
    # достаем Аптека
    print_values("C13", 20, input_ws, output_ws, rows_indent, filename)

    rows_indent += 1
try:
    if os.path.exists(output_path):
        print(f"Файл уже существует, перезаписываем: {output_path}")
    output_wb.save(output_path)
    print(f"Файл успешно сохранен! {output_path}")
except Exception as e:
    print(f"[ОШИБКА] Не удалось сохранить файл: {e}")



